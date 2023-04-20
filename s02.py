import openpyxl
import pathlib
import shutil

path_0 = pathlib.Path(__file__).parent
pathname_1=path_0.joinpath("test2.xlsx").resolve()
pathname_2=path_0.joinpath("test3.xlsx").resolve()
pathname_3=path_0.joinpath("test2_copy.xlsx").resolve()

## 方法１ ##

# ws1,ws2を取得
src_wb = openpyxl.load_workbook(pathname_1)
ws1 = src_wb["ws1"]
ws2 = src_wb["ws2"]
src_wb.close()

# ws1,ws2を新bookへsave → 問題のあるブックができる
dst_wb = openpyxl.Workbook()
dst_wb._sheets.append(ws1)
dst_wb._sheets.append(ws2)
dst_wb.save(pathname_2)

## 方法２ ##

# test2.xlsx をコピー
shutil.copyfile(pathname_1,pathname_3)

# ws1,ws2を取得
src_wb = openpyxl.load_workbook(pathname_3)
ws1 = src_wb["ws1"]
ws2 = src_wb["ws2"]

# ws1,ws2の内容を変更
ws1['B1'].value = 3
ws2['A2'].value = 7
ws2['B2'].value = 8
ws2['C2'].value = 9

# ws1,ws2を新bookへsave → OK
src_wb.save(pathname_3)

